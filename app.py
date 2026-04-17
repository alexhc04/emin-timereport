"""
EMIN TIME REPORT - Aplicación de Control Horario
================================================
Despacho de Abogados Emin
Backend: Flask (Python)
Base de datos: PostgreSQL (Supabase)
"""

from flask import Flask, render_template, request, jsonify, session, redirect, url_for, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
from datetime import datetime, date
import os
from functools import wraps

# ─────────────────────────────────────────────
# CONFIGURACIÓN
# ─────────────────────────────────────────────
app = Flask(__name__)
app.secret_key = os.environ.get('SECRET_KEY', 'emin-secret-key-2024')

DATABASE_URL = os.environ.get('DATABASE_URL', 'sqlite:///emin_local.db')
if DATABASE_URL.startswith("postgres://"):
    DATABASE_URL = DATABASE_URL.replace("postgres://", "postgresql://", 1)

app.config['SQLALCHEMY_DATABASE_URI'] = DATABASE_URL
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db = SQLAlchemy(app)

# ─────────────────────────────────────────────
# LISTA CERRADA DE TIPOS DE TAREA
# ─────────────────────────────────────────────
TIPOS_TAREA = [
    ('REUNION INTERNA/FORMACION',      False),   # no productiva
    ('DESCANSO',                       False),   # no productiva
    ('REUNION CON CLIENTE/CONTRARIO',  True),
    ('LLAMADA A CLIENTE/CONTRARIO',    True),
    ('EMAIL A CLIENTE/CONTRARIO',      True),
    ('ANAL. VIABILIDAD/INFORME',       True),
    ('ASISTENCIA A FIRMA',             True),
    ('PREP. CONTRATO/ACTA',            True),
    ('REDACCION DEMANDA/RECURSO',      True),
    ('ASISTENCIA JUICIO/CONCILIACION', True),
    ('TRAMITES ADMON PUBLICA',         True),
]
TAREAS_IMPRODUCTIVAS = [t[0] for t in TIPOS_TAREA if not t[1]]


# ─────────────────────────────────────────────
# MODELOS
# ─────────────────────────────────────────────

class Usuario(db.Model):
    __tablename__ = 'usuarios'
    id            = db.Column(db.Integer, primary_key=True)
    nombre        = db.Column(db.String(100), nullable=False)
    username      = db.Column(db.String(50), unique=True, nullable=False)
    password_hash = db.Column(db.String(255), nullable=False)
    rol           = db.Column(db.String(20), nullable=False, default='worker')
    activo        = db.Column(db.Boolean, default=True)
    creado        = db.Column(db.DateTime, default=datetime.utcnow)
    tareas        = db.relationship('Tarea', backref='usuario', lazy=True)

    def set_password(self, password):
        self.password_hash = generate_password_hash(password)

    def check_password(self, password):
        return check_password_hash(self.password_hash, password)

    def to_dict(self):
        return {'id': self.id, 'nombre': self.nombre,
                'username': self.username, 'rol': self.rol}


class Tarea(db.Model):
    __tablename__ = 'tareas'
    id          = db.Column(db.Integer, primary_key=True)
    usuario_id  = db.Column(db.Integer, db.ForeignKey('usuarios.id'), nullable=False)
    tipo        = db.Column(db.String(60),  nullable=False)
    hora_inicio = db.Column(db.Time,        nullable=False)
    hora_fin    = db.Column(db.Time,        nullable=False)
    cliente     = db.Column(db.String(150))
    dni_nif     = db.Column(db.String(20))
    ref_bitrix  = db.Column(db.String(100))
    contacto    = db.Column(db.String(150))
    fecha       = db.Column(db.Date, nullable=False, default=date.today)
    notas       = db.Column(db.String(30))
    creado      = db.Column(db.DateTime, default=datetime.utcnow)

    def duracion_horas(self):
        inicio = datetime.combine(date.today(), self.hora_inicio)
        fin    = datetime.combine(date.today(), self.hora_fin)
        return round((fin - inicio).total_seconds() / 3600, 2)

    def es_productiva(self):
        return self.tipo not in TAREAS_IMPRODUCTIVAS

    def to_dict(self):
        return {
            'id':             self.id,
            'usuario_id':     self.usuario_id,
            'usuario_nombre': self.usuario.nombre if self.usuario else '',
            'tipo':           self.tipo,
            'hora_inicio':    self.hora_inicio.strftime('%H:%M'),
            'hora_fin':       self.hora_fin.strftime('%H:%M'),
            'cliente':        self.cliente    or '',
            'dni_nif':        self.dni_nif    or '',
            'ref_bitrix':     self.ref_bitrix or '',
            'contacto':       self.contacto   or '',
            'fecha':          self.fecha.strftime('%Y-%m-%d'),
            'notas':          self.notas      or '',
            'duracion':       self.duracion_horas(),
            'productiva':     self.es_productiva(),
        }


# ─────────────────────────────────────────────
# DECORADORES
# ─────────────────────────────────────────────

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if 'user_id' not in session:
            return redirect(url_for('login'))
        return f(*args, **kwargs)
    return decorated


# ─────────────────────────────────────────────
# PÁGINAS
# ─────────────────────────────────────────────

@app.route('/')
def index():
    return redirect(url_for('dashboard') if 'user_id' in session else url_for('login'))


@app.route('/login', methods=['GET', 'POST'])
def login():
    if request.method == 'GET':
        return redirect(url_for('dashboard')) if 'user_id' in session else render_template('login.html')

    data     = request.get_json()
    username = data.get('username', '').strip()
    password = data.get('password', '')
    usuario  = Usuario.query.filter_by(username=username, activo=True).first()

    if usuario and usuario.check_password(password):
        session['user_id']  = usuario.id
        session['username'] = usuario.username
        session['nombre']   = usuario.nombre
        session['rol']      = usuario.rol
        return jsonify({'success': True, 'rol': usuario.rol})
    return jsonify({'success': False, 'error': 'Usuario o contraseña incorrectos'}), 401


@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))


@app.route('/dashboard')
@login_required
def dashboard():
    return render_template('dashboard.html',
                           nombre=session['nombre'],
                           rol=session['rol'],
                           username=session['username'])


# ─────────────────────────────────────────────
# API: TAREAS
# ─────────────────────────────────────────────

@app.route('/api/tareas', methods=['GET'])
@login_required
def get_tareas():
    user_id = session['user_id']
    rol     = session['rol']

    filtro_usuario = request.args.get('usuario_id')
    filtro_dni     = request.args.get('dni_nif',    '').strip()
    filtro_bitrix  = request.args.get('ref_bitrix', '').strip()
    filtro_desde   = request.args.get('fecha_desde')
    filtro_hasta   = request.args.get('fecha_hasta')

    query = Tarea.query

    if rol == 'worker':
        query = query.filter_by(usuario_id=user_id)
    elif rol == 'admin' and filtro_usuario:
        query = query.filter_by(usuario_id=filtro_usuario)

    if filtro_dni:
        query = query.filter(Tarea.dni_nif.ilike(f'%{filtro_dni}%'))
    if filtro_bitrix:
        query = query.filter(Tarea.ref_bitrix.ilike(f'%{filtro_bitrix}%'))
    if filtro_desde:
        query = query.filter(Tarea.fecha >= filtro_desde)
    if filtro_hasta:
        query = query.filter(Tarea.fecha <= filtro_hasta)

    tareas = query.order_by(Tarea.fecha.desc(), Tarea.creado.desc()).all()
    return jsonify([t.to_dict() for t in tareas])


@app.route('/api/tareas', methods=['POST'])
@login_required
def crear_tarea():
    data = request.get_json()

    # Campos siempre obligatorios
    for campo in ['tipo', 'hora_inicio', 'hora_fin']:
        if not data.get(campo):
            return jsonify({'error': f'El campo {campo} es obligatorio'}), 400

    # Validar tipo de tarea
    tipos_validos = [t[0] for t in TIPOS_TAREA]
    if data['tipo'] not in tipos_validos:
        return jsonify({'error': 'Tipo de tarea no válido'}), 400

    # Campos obligatorios para tareas productivas
    if data['tipo'] not in TAREAS_IMPRODUCTIVAS:
        for campo, nombre in [('cliente', 'Nombre cliente'),
                              ('dni_nif', 'DNI/NIF/NIE'),
                              ('ref_bitrix', 'Ref. caso Bitrix'),
                              ('contacto', 'Persona de contacto')]:
            if not data.get(campo, '').strip():
                return jsonify({'error': f'{nombre} es obligatorio'}), 400

    # Validar horas
    try:
        hora_inicio = datetime.strptime(data['hora_inicio'], '%H:%M').time()
        hora_fin    = datetime.strptime(data['hora_fin'],    '%H:%M').time()
    except ValueError:
        return jsonify({'error': 'Formato de hora inválido (HH:MM)'}), 400

    if hora_fin <= hora_inicio:
        return jsonify({'error': 'La hora de fin debe ser posterior al inicio'}), 400

    # Validar fecha
    try:
        fecha = datetime.strptime(
            data.get('fecha', date.today().strftime('%Y-%m-%d')), '%Y-%m-%d').date()
    except ValueError:
        return jsonify({'error': 'Formato de fecha inválido'}), 400

    # Validar longitud del asunto
    notas = data.get('notas', '')
    if len(notas) > 30:
        return jsonify({'error': 'El asunto no puede superar 30 caracteres'}), 400

    tarea = Tarea(
        usuario_id  = session['user_id'],
        tipo        = data['tipo'],
        hora_inicio = hora_inicio,
        hora_fin    = hora_fin,
        cliente     = data.get('cliente',    '').strip(),
        dni_nif     = data.get('dni_nif',    '').strip().upper(),
        ref_bitrix  = data.get('ref_bitrix', '').strip(),
        contacto    = data.get('contacto',   '').strip(),
        fecha       = fecha,
        notas       = notas[:30],
    )
    db.session.add(tarea)
    db.session.commit()
    return jsonify({'success': True, 'tarea': tarea.to_dict()}), 201


@app.route('/api/tareas/<int:tarea_id>', methods=['DELETE'])
@login_required
def eliminar_tarea(tarea_id):
    tarea = Tarea.query.get_or_404(tarea_id)
    if tarea.usuario_id != session['user_id'] and session['rol'] != 'admin':
        return jsonify({'error': 'Sin permiso'}), 403
    db.session.delete(tarea)
    db.session.commit()
    return jsonify({'success': True})


@app.route('/api/resumen', methods=['GET'])
@login_required
def get_resumen():
    user_id = session['user_id']
    rol     = session['rol']

    filtro_usuario = request.args.get('usuario_id')
    filtro_desde   = request.args.get('fecha_desde')
    filtro_hasta   = request.args.get('fecha_hasta')

    query = Tarea.query
    if rol == 'worker':
        query = query.filter_by(usuario_id=user_id)
    elif rol == 'admin' and filtro_usuario:
        query = query.filter_by(usuario_id=filtro_usuario)
    if filtro_desde:
        query = query.filter(Tarea.fecha >= filtro_desde)
    if filtro_hasta:
        query = query.filter(Tarea.fecha <= filtro_hasta)

    tareas = query.all()
    prod   = sum(t.duracion_horas() for t in tareas if     t.es_productiva())
    improd = sum(t.duracion_horas() for t in tareas if not t.es_productiva())
    return jsonify({
        'horas_productivas':   round(prod,   2),
        'horas_improductivas': round(improd, 2),
        'total_tareas':        len(tareas),
    })


# ─────────────────────────────────────────────
# API: USUARIOS
# ─────────────────────────────────────────────

@app.route('/api/usuarios', methods=['GET'])
@login_required
def get_usuarios():
    if session['rol'] == 'admin':
        return jsonify([u.to_dict() for u in Usuario.query.filter_by(activo=True).all()])
    return jsonify([Usuario.query.get(session['user_id']).to_dict()])


# ─────────────────────────────────────────────
# API: EXPORTAR EXCEL
# ─────────────────────────────────────────────

@app.route('/api/exportar', methods=['GET'])
@login_required
def exportar_excel():
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import io

    user_id = session['user_id']
    rol     = session['rol']

    filtro_usuario = request.args.get('usuario_id')
    filtro_dni     = request.args.get('dni_nif',    '').strip()
    filtro_bitrix  = request.args.get('ref_bitrix', '').strip()
    filtro_desde   = request.args.get('fecha_desde')
    filtro_hasta   = request.args.get('fecha_hasta')

    query = Tarea.query
    if rol == 'worker':
        query = query.filter_by(usuario_id=user_id)
    elif rol == 'admin' and filtro_usuario:
        query = query.filter_by(usuario_id=filtro_usuario)
    if filtro_dni:
        query = query.filter(Tarea.dni_nif.ilike(f'%{filtro_dni}%'))
    if filtro_bitrix:
        query = query.filter(Tarea.ref_bitrix.ilike(f'%{filtro_bitrix}%'))
    if filtro_desde:
        query = query.filter(Tarea.fecha >= filtro_desde)
    if filtro_hasta:
        query = query.filter(Tarea.fecha <= filtro_hasta)

    tareas = query.order_by(Tarea.fecha.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Informe de Horas"

    navy_hex  = "0D1B2A"
    green_hex = "D4EDDA"
    amber_hex = "FFF3CD"

    hfont = Font(bold=True, color="FFFFFF", size=10)
    hfill = PatternFill(start_color=navy_hex, end_color=navy_hex, fill_type="solid")
    halign = Alignment(horizontal="center", vertical="center")
    borde  = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'),  bottom=Side(style='thin'))

    # Título
    ws.merge_cells('A1:L1')
    ws['A1'] = "INFORME DE HORAS — EMIN ABOGADOS"
    ws['A1'].font = Font(bold=True, size=13, color=navy_hex)
    ws['A1'].alignment = Alignment(horizontal="center")

    ws.merge_cells('A2:L2')
    ws['A2'] = f"Generado el {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A2'].alignment = Alignment(horizontal="center")
    ws['A2'].font = Font(size=9, color="777777")

    # Cabeceras
    headers = ['Fecha','Usuario','Tipo de Tarea','Inicio','Fin',
               'Duración (h)','Cliente','DNI/NIF/NIE','Ref. Bitrix',
               'Contacto','Asunto','Productiva']
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=4, column=col, value=h)
        c.font = hfont; c.fill = hfill
        c.alignment = halign; c.border = borde

    # Filas
    for ri, t in enumerate(tareas, 5):
        color = green_hex if t.es_productiva() else amber_hex
        rfill = PatternFill(start_color=color, end_color=color, fill_type="solid")
        vals  = [
            t.fecha.strftime('%d/%m/%Y'),
            t.usuario.nombre,
            t.tipo,
            t.hora_inicio.strftime('%H:%M'),
            t.hora_fin.strftime('%H:%M'),
            t.duracion_horas(),
            t.cliente    or '',
            t.dni_nif    or '',
            t.ref_bitrix or '',
            t.contacto   or '',
            t.notas      or '',
            'Sí' if t.es_productiva() else 'No',
        ]
        for ci, v in enumerate(vals, 1):
            c = ws.cell(row=ri, column=ci, value=v)
            c.fill = rfill; c.border = borde
            c.alignment = Alignment(vertical="center")

    # Totales
    tr = len(tareas) + 6
    ws.cell(row=tr, column=1, value="TOTALES").font = Font(bold=True)
    ws.cell(row=tr, column=6,
            value=round(sum(t.duracion_horas() for t in tareas), 2)).font = Font(bold=True)

    # Anchos
    for i, w in enumerate([12,20,30,8,8,13,25,16,18,22,20,11], 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    fname = f"emin_informe_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name=fname)


# ─────────────────────────────────────────────
# DIAGNÓSTICO
# ─────────────────────────────────────────────

@app.route('/health')
def health():
    info = {
        'python':               __import__('sys').version,
        'database_url_set':     bool(os.environ.get('DATABASE_URL')),
    }
    try:
        info['usuarios'] = Usuario.query.count()
        info['status']   = 'ok'
    except Exception as e:
        info['status'] = 'error_db'
        info['error']  = str(e)
    return jsonify(info)


# Ruta para migrar la base de datos (añadir columnas nuevas si no existen)
@app.route('/migrate-db-emin-interno')
def migrate_db():
    try:
        with db.engine.connect() as conn:
            # Intentar añadir cada columna nueva — si ya existe, el error se ignora
            for sql in [
                "ALTER TABLE tareas ADD COLUMN IF NOT EXISTS ref_bitrix VARCHAR(100)",
                "ALTER TABLE tareas ADD COLUMN IF NOT EXISTS contacto VARCHAR(150)",
                "ALTER TABLE tareas ALTER COLUMN notas TYPE VARCHAR(30)",
            ]:
                try:
                    conn.execute(db.text(sql))
                    conn.commit()
                except Exception as col_err:
                    pass  # La columna ya existía, no pasa nada
        return jsonify({'ok': True, 'mensaje': 'Migración completada correctamente'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# Ruta temporal para actualizar contraseñas — úsala una vez y luego elimínala
@app.route('/reset-passwords-emin-interno')
def reset_passwords():
    try:
        cambios = [
            ('admin',       'Rm7vK2nX9p'),
            ('trabajador1', 'Hj4wQ8tL3m'),
            ('trabajador2', 'Zc6bN1sP5k'),
        ]
        for username, pwd in cambios:
            u = Usuario.query.filter_by(username=username).first()
            if u:
                u.set_password(pwd)
        db.session.commit()
        return jsonify({'ok': True, 'mensaje': 'Contraseñas actualizadas. Elimina esta ruta del código.'})
    except Exception as e:
        return jsonify({'ok': False, 'error': str(e)}), 500


# ─────────────────────────────────────────────
# INICIALIZACIÓN
# ─────────────────────────────────────────────

def init_db():
    with app.app_context():
        db.create_all()
        if Usuario.query.count() == 0:
            for datos in [
                ('Administrador CEO', 'admin',       'Rm7vK2nX9p', 'admin'),
                ('Trabajador Uno',    'trabajador1', 'Hj4wQ8tL3m', 'worker'),
                ('Trabajador Dos',    'trabajador2', 'Zc6bN1sP5k', 'worker'),
            ]:
                u = Usuario(nombre=datos[0], username=datos[1], rol=datos[3])
                u.set_password(datos[2])
                db.session.add(u)
            db.session.commit()
            print("✅ Usuarios creados con nuevas contraseñas")
        else:
            print("ℹ️  Usuarios ya existen en la base de datos")


try:
    init_db()
except Exception as e:
    print(f"⚠️  init_db() falló: {e}")

if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
